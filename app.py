import streamlit as st
import pandas as pd
import openpyxl
from openpyxl import load_workbook
from openpyxl.styles import Border, Side
from openpyxl.utils import get_column_letter

PASSWORD = "IMR Creation"

# Function to process country data with given base year, base value, periods, and YoY changes
def process_country_data(base_year, base_value, periods, yoy_changes, is_future=True):
    data = []
    value = base_value
    if is_future:
        for period, yoy_change in zip(periods, yoy_changes):
            value *= (1 + yoy_change)
            data.append((period, value))
    else:
        for period, yoy_change in zip(periods, yoy_changes):
            value /= (1 + yoy_change)
            data.append((period, value))
    return data

# Function to calculate region data based on base model data, region shares, region YoY changes, base year, and base value
def calculate_region_data(base_model_data, region_shares, region_yoy_changes, base_year, base_value):
    region_data = []
    region_values = {region: base_value * share / 100 for region, share in region_shares}
    initial_region_values = region_values.copy()
 
    for year, total_value in base_model_data:
        regions = {}
        if year == base_year:
            regions = initial_region_values.copy()
        else:
            for region in initial_region_values:
                if year > base_year:
                    region_values[region] *= (1 + region_yoy_changes[year].get(region, 0))
                else:
                    region_values[region] /= (1 + region_yoy_changes[year].get(region, 0))
                regions[region] = region_values[region]
 
            total_region_value = sum(regions.values())
            adjustment_factor = total_value / total_region_value if total_region_value != 0 else 1
            for region in regions:
                regions[region] *= adjustment_factor
 
        region_data.append((year, regions))
    return region_data

#  Function to calculate country data based on region data, country shares, country YoY changes, and base year
def calculate_country_data(region_data, country_shares, country_yoy_changes, base_year):
    country_data = []
    for year, regions in region_data:
        all_countries = {}
        for region, value in regions.items():
            if region in country_shares:
                country_values = {country: value * share / 100 for country, share in country_shares[region].items()}
                if year != base_year:
                    for country in country_values:
                        if year > base_year:
                            country_values[country] *= (1 + country_yoy_changes[year][region].get(country, 0))
                        else:
                            country_values[country] /= (1 + country_yoy_changes[year][region].get(country, 0))
                    total_country_value = sum(country_values.values())
                    adjustment_factor = value / total_country_value if total_country_value != 0 else 1
                    for country in country_values:
                        country_values[country] *= adjustment_factor
                all_countries[region] = country_values
        country_data.append((year, all_countries))
    return country_data

# Function to calculate segment data based on country data, segments, segment shares, segment YoY changes, base year
def calculate_segment_data(country_data, segment_shares, segment_yoy_changes, base_year):
    segment_data = []
    for year, regions in country_data:
        all_segments = {}
        for region, countries in regions.items():
            segment_values = {}
            for country, value in countries.items():
                if country in segment_shares[region]:
                    # Assign country value to each segment directly
                    segment_values[country] = {segment: value for segment in segment_shares[region][country]}
            all_segments[region] = segment_values
        segment_data.append((year, all_segments))
    return segment_data


# Function to calculate subsegment data based on segment data, subsegments, subsegment shares, subsegment YoY changes, and base year
def calculate_subsegment_data(segment_data, subsegments, subsegment_shares, subsegment_yoy_changes, base_year):
    subsegment_data = []
    for year, regions in segment_data:
        all_subsegments = {}
        for region, countries in regions.items():
            subsegment_values = {}
            for country, segments in countries.items():
                subsegment_values[country] = {}
                for segment, value in segments.items():
                    if segment in subsegment_shares[region][country]:
                        initial_subsegment_values = {subsegment: value * share / 100 for subsegment, share in subsegment_shares[region][country][segment].items()}
                        subsegment_values[country][segment] = initial_subsegment_values
                        if year != base_year:
                            for subsegment in initial_subsegment_values:
                                if year > base_year:
                                    initial_subsegment_values[subsegment] *= (1 + subsegment_yoy_changes[year][region][country][segment].get(subsegment, 0))
                                else:
                                    initial_subsegment_values[subsegment] /= (1 + subsegment_yoy_changes[year][region][country][segment].get(subsegment, 0))
                            total_subsegment_value = sum(initial_subsegment_values.values())
                            adjustment_factor = value / total_subsegment_value if total_subsegment_value != 0 else 1
                            for subsegment in initial_subsegment_values:
                                initial_subsegment_values[subsegment] *= adjustment_factor
            all_subsegments[region] = subsegment_values
        subsegment_data.append((year, all_subsegments))
    return subsegment_data

from math import pow
# def shift_data_up(ws):
#     """Remove blank rows by shifting the data upwards."""
#     max_row = ws.max_row
#     max_col = ws.max_column

#     for row in range(1, max_row + 1):
#         if all(ws.cell(row=row, column=col).value is None for col in range(1, max_col + 1)):
#             ws.delete_rows(row)

def shift_data_left(ws):
    """Remove blank columns by shifting the data to the left."""
    max_row = ws.max_row
    max_col = ws.max_column

    for col in range(1, max_col + 1):
        if all(ws.cell(row=row, column=col).value is None for row in range(1, max_row + 1)):
            ws.delete_cols(col)

def add_total_row(ws, start_row, end_row, max_col_used, table_name_row, is_first_table, first_table_total_row_data, is_region_table, copy_first_row=False):
    total_row = end_row + 1
    start_col = 1
    end_col = max_col_used

    # Add total row label
    ws.cell(row=total_row, column=start_col, value="Total")

    if is_first_table:
        # For the first table, sum the columns as usual and store the total row data
        for col in range(start_col + 1, end_col + 1):
            col_letter = get_column_letter(col)
            sum_formula = f"=SUM({col_letter}{start_row}:{col_letter}{end_row})"
            ws.cell(row=total_row, column=col, value=sum_formula)

        # Store the total row data for later use in region tables
        first_table_total_row_data.clear()
        first_table_total_row_data.extend([ws.cell(row=total_row, column=col).value for col in range(start_col + 1, end_col + 1)])
        
    else:
        # For region tables, replace the second row with the total row from the first table
        if is_region_table:
            for col in range(start_col + 1, end_col + 1):
                value_to_copy = first_table_total_row_data[col - 2]  # Copy from the first table's total row (adjust index for correct column)
                ws.cell(row=start_row + 1, column=col, value=value_to_copy)

    # Copy the value from the first row to the total row if required
    if copy_first_row:
        for col in range(start_col + 1, end_col + 1):
            first_row_value = ws.cell(row=start_row + 1, column=col).value
            ws.cell(row=total_row, column=col, value=first_row_value)

    # Add CAGR column next to the last column without gaps, aligned with the table name row
    cagr_col = end_col + 1
    ws.cell(row=table_name_row, column=cagr_col, value="CAGR")  # Add "CAGR" to the same row as the table name

    for row in range(start_row + 1, total_row + 1):  # Ensure CAGR calculation is done for all rows above the total row
        start_value = ws.cell(row=row, column=start_col + 1).value
        end_value = ws.cell(row=row, column=end_col).value
        if start_value and end_value and start_value != 0:
            periods = (end_col - (start_col + 1)) + 1  # Number of periods between the second column and the last column
            if periods > 0:
                cagr_formula = f"=({get_column_letter(end_col)}{row}/{get_column_letter(start_col + 1)}{row})^(1/{periods})-1"
                cell = ws.cell(row=row, column=cagr_col, value=cagr_formula)
                cell.number_format = '0.00%'  # Format as a percentage with two decimal places
            else:
                ws.cell(row=row, column=cagr_col, value=None)  # Not enough periods for CAGR calculation
        else:
            ws.cell(row=row, column=cagr_col, value=None)  # Leave blank if the data is not available



    # Add borders to the total row and the CAGR column
    thin_border = Border(left=Side(style='thin'), 
                         right=Side(style='thin'), 
                         top=Side(style='thin'), 
                         bottom=Side(style='thin'))
    for col in range(start_col, cagr_col + 1):  # Include the CAGR column in the border range
        ws.cell(row=total_row, column=col).border = thin_border

    for row in range(start_row, total_row + 1):
        ws.cell(row=row, column=cagr_col).border = thin_border
    return cagr_col  # Return the new maximum column used


def add_borders_to_table(ws, start_row, end_row, start_col, end_col):
    thin_border = Border(left=Side(style='thin'), 
                         right=Side(style='thin'), 
                         top=Side(style='thin'), 
                         bottom=Side(style='thin'))

    for row in ws.iter_rows(min_row=start_row, max_row=end_row, min_col=start_col, max_col=end_col):
        for cell in row:
            cell.border = thin_border


def format_excel_sheet(ws):
    # Step 1: Remove blank rows and columns to eliminate gaps
    # shift_data_up(ws)
    shift_data_left(ws)

    # Step 2: Identify and add total rows for each table
    start_row = 2
    max_col_used = ws.max_column  # Track the maximum column used so far
    table_name_row = start_row  # Initialize the table name row
    is_first_table = True  # Start with the assumption that we are processing the first table
    first_table_total_row_data = []  # To store the total row data of the first table

    for row in ws.iter_rows(min_row=2, max_col=max_col_used):
        if all(cell.value is None for cell in row):  # If a blank row is found
            end_row = row[0].row - 1
            if start_row <= end_row:
                # Determine if the current table is a region table
                table_name = ws.cell(row=table_name_row, column=1).value
                sheet_name = ws.title
                is_region_table = table_name and table_name.startswith(sheet_name)

                add_total_row(ws, start_row, end_row, max_col_used, table_name_row, is_first_table, first_table_total_row_data, is_region_table, copy_first_row=not is_first_table)
                add_borders_to_table(ws, start_row, end_row, 1, max_col_used)
                is_first_table = False  # After the first table is processed, this flag is set to False
            start_row = row[0].row + 1
            table_name_row = start_row  # Update the table name row

    # Add total row for the last table if not followed by a blank row
    end_row = ws.max_row
    if start_row <= end_row:
        # Determine if the last table is a region table
        table_name = ws.cell(row=table_name_row, column=1).value
        sheet_name = ws.title
        is_region_table = table_name and table_name.startswith(sheet_name)

        add_total_row(ws, start_row, end_row, max_col_used, table_name_row, is_first_table, first_table_total_row_data, is_region_table, copy_first_row=not is_first_table)
        add_borders_to_table(ws, start_row, end_row, 1, max_col_used)

import io

def save_data_to_excel(base_model_data, region_data, country_data, segment_data, subsegment_data):
    output = io.BytesIO()  # Create a BytesIO buffer to hold the Excel data

    try:
        # Convert base model data to DataFrame
        base_model_df = pd.DataFrame(base_model_data, columns=['Year', 'Value'])

        # Convert and transpose region data to have regions as rows and years as columns
        region_list = {}
        for year, data in region_data:
            region_list[year] = data
        region_df = pd.DataFrame(region_list)
        region_df.index.name = 'Region'

        # Create an Excel writer object using the buffer
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            region_df.to_excel(writer, sheet_name='Global Data', index_label='Region/year')

            for region in country_data[0][1]:
                combined_country_data = {}
                for year, data in country_data:
                    if year not in combined_country_data:
                        combined_country_data[year] = data[region]

                combined_country_df = pd.DataFrame(combined_country_data)
                combined_country_df.index.name = 'Country'
                combined_country_df.to_excel(writer, sheet_name=f'{region}', index_label='Country')

                startrow = len(combined_country_df) + 3  # Leave a gap of one line

                for country in combined_country_df.index:
                    for segment in segment_data[0][1][region][country]:
                        segment_data_dict = {}
                        for year, data in segment_data:
                            if country in data[region] and segment in data[region][country]:
                                segment_data_dict[year] = {segment: data[region][country][segment]}

                        for year, data in subsegment_data:
                            if country in data[region] and segment in data[region][country]:
                                for subsegment, value in data[region][country][segment].items():
                                    if year in segment_data_dict:
                                        segment_data_dict[year].update({subsegment: value})
                                    else:
                                        segment_data_dict[year] = {subsegment: value}

                        segment_df = pd.DataFrame(segment_data_dict)
                        segment_df['Country'] = country
                        segment_df = segment_df.reset_index().rename(columns={'index': 'Year'})
                        segment_df.columns = [f'{country}/Year' if col == 'Year' else col for col in segment_df.columns]
                        segment_df.drop(columns=['Country'], inplace=True)
                        segment_df.to_excel(writer, sheet_name=f'{region}', startrow=startrow, index=False)
                        startrow += len(segment_df) + 3  # Leave a gap of one line

                region_subsegment_data = {}
                for year, data in subsegment_data:
                    for country in data[region]:
                        for segment in data[region][country]:
                            for subsegment, value in data[region][country][segment].items():
                                if year not in region_subsegment_data:
                                    region_subsegment_data[year] = {}
                                if segment not in region_subsegment_data[year]:
                                    region_subsegment_data[year][segment] = {}
                                if subsegment not in region_subsegment_data[year][segment]:
                                    region_subsegment_data[year][segment][subsegment] = 0
                                region_subsegment_data[year][segment][subsegment] += value

                is_first_table = True

                for segment in region_subsegment_data[year]:
                    segment_subsegment_data = {}
                    segment_total_values = {}

                    for year in region_subsegment_data:
                        subsegment_values = region_subsegment_data[year][segment]

                        subsegment_values_filtered = {key: value for key, value in subsegment_values.items() if key != 'year'}

                        if is_first_table:
                            segment_total_values[year] = sum(value for value in subsegment_values_filtered.values() if isinstance(value, (int, float)))
                        else:
                            if len(subsegment_values_filtered) > 1:
                                second_row_key = list(subsegment_values_filtered.keys())[1]
                                segment_total_values[year] = subsegment_values_filtered[second_row_key]
                            else:
                                segment_total_values[year] = None

                        segment_subsegment_data[year] = subsegment_values_filtered

                    segment_subsegment_df = pd.DataFrame(segment_subsegment_data)
                    segment_subsegment_df.index.name = f'{region}/year'

                    segment_total_row = pd.DataFrame(segment_total_values, index=[f'{segment}'])

                    segment_subsegment_df_with_total = pd.concat([segment_total_row, segment_subsegment_df])

                    segment_subsegment_df_with_total.to_excel(writer, sheet_name=f'{region}', startrow=startrow, index_label=f'{region}/year')
                    startrow += len(segment_subsegment_df_with_total) + 3

                    is_first_table = False
             
        wb = load_workbook(output)
        for sheetname in wb.sheetnames:
            ws = wb[sheetname]
            format_excel_sheet(ws)

        wb.save(output)
        print("Data saved successfully!")
        return output  # Return the BytesIO buffer instead of saving to file

    except PermissionError as e:
        st.error(f"PermissionError: {e}. Please make sure the file is not open or you have the correct permissions.")
    except Exception as e:
        st.error(f"An error occurred: {e}")
        return None

    

def main():
    st.title("Market Data Projection Tool")
    
    if 'data' not in st.session_state:
        st.session_state.data = {}
        st.session_state.page = 'Input Data'
    
    page = st.sidebar.radio("Navigation", ["Input Data", "Review Data"])
    
    if page == "Input Data":
        st.header("Input Data")

        # Group inputs into a form
        with st.form(key='data_form'):
            base_year = st.number_input("Enter the base year:", min_value=1900, max_value=2100, value=2023, step=1)
            base_value = st.number_input("Enter the value for the base year:", min_value=0.0, value=1000.0, step=0.01)

            future_periods_count = st.number_input("Enter the number of future periods to predict:", min_value=0, max_value=100, value=1, step=1)
            future_periods = [base_year + i + 1 for i in range(future_periods_count)]
            future_yoy_changes = [st.number_input(f"Enter the YoY change for year {year}:", min_value=-1.000, max_value=2.000, value=0.065, step=0.001, format="%.4f") for year in future_periods]

            previous_periods_count = st.number_input("Enter the number of previous periods to predict:", min_value=0, max_value=100, value=1, step=1)
            previous_periods = [base_year - i - 1 for i in range(previous_periods_count)]
            previous_yoy_changes = [st.number_input(f"Enter the YoY change for year {year}:", min_value=-1.000, max_value=1.000, value=0.050, step=0.001, format="%.4f") for year in previous_periods]

            regions = ["North America", "Eastern Europe", 'Western Europe', 'Asia Pacific', 'Middle East and Africa', 'South America']
            # regions = ["North America", "Eastern Europe"]
            region_shares = [(region, st.number_input(f"Enter the percentage share of {region}:", min_value=0.0, max_value=100.0, step=0.1)) for region in regions]

            region_yoy_changes = {}
            for period in previous_periods + future_periods:
                region_yoy_changes[period] = {region: st.number_input(f"Enter the YoY change for {region} in year {period}:", min_value=-1.000, max_value=2.000, value=0.065, step=0.001, format="%.4f") for region in regions}

            countries = {
                "North America": ["US", "Canada", "Mexico"],
                "Eastern Europe": ["Bulgaria", "Czech Republic", "Hungary", "Poland", "Romania", "Russia", "Rest of Eastern Europe"],
                "Western Europe": ["Germany", "UK", "France", "Netherlands", "Italy", "Spain", "Rest of Western Europe"],
                "Asia Pacific": ["China", "India", "Japan", "South Korea", "Malaysia", "Thailand", "Vietnam", "Philippines", "Australia", "New Zealand", "Rest of Asia Pacific"],
                "Middle East and Africa": ["Turkey", "Bahrain", "Kuwait", "Saudi Arabia", "Qatar", "UAE", "Israel", "South Africa"],
                "South America": ["Brazil", "Argentina", "Rest of South America"]
            }

            country_shares = {}
            for region in countries:
                country_shares[region] = {}
                for country in countries[region]:
                    country_shares[region][country] = st.number_input(f"Enter the percentage share of {country} in {region}:", min_value=0.0, max_value=100.0, step=0.1)

            country_yoy_changes = {}
            for period in previous_periods + future_periods:
                country_yoy_changes[period] = {}
                for region in countries:
                    country_yoy_changes[period][region] = {}
                    for country in countries[region]:
                        country_yoy_changes[period][region][country] = st.number_input(f"Enter the YoY change for {country} in {region} in year {period}:", min_value=-1.000, max_value=2.000, value=0.065, step=0.001, format="%.4f")

            # New dynamic input for segment and subsegment count
            segments_count = st.number_input("Enter the number of segments:", min_value=1, max_value=100, value=2, step=1)
            segments = [st.text_input(f"Enter the name of segment {i + 1}:", value=f"Segment {i + 1}") for i in range(segments_count)]

            subsegments = {}
            for segment in segments:
                subsegments_count = st.number_input(f"Enter the number of subsegments in {segment}:", min_value=1, max_value=100, value=2, step=1)
                subsegments[segment] = [st.text_input(f"Enter the name of subsegment {j + 1} in {segment}:", value=f"{segment}_Subsegment_{j + 1}") for j in range(subsegments_count)]

            segment_shares = {}
            for region in regions:
                segment_shares[region] = {}
                for country in countries[region]:
                    segment_shares[region][country] = {segment: st.number_input(f"Enter the share of {segment} in {country} ({region}):", min_value=0.0, max_value=100.0, value=100.0, step=0.1) for segment in segments}

            subsegment_shares = {}
            for region in regions:
                subsegment_shares[region] = {}
                for country in countries[region]:
                    subsegment_shares[region][country] = {}
                    for segment in segments:
                        subsegment_shares[region][country][segment] = {subsegment: st.number_input(f"Enter the share of {subsegment} in {segment} ({country}, {region}):", min_value=0.0, max_value=100.0, value=100.0 / len(subsegments[segment]), step=0.1) for subsegment in subsegments[segment]}

            segment_yoy_changes = {}
            for period in previous_periods + future_periods:
                segment_yoy_changes[period] = {}
                for region in countries:
                    segment_yoy_changes[period][region] = {}
                    for country in countries[region]:
                        segment_yoy_changes[period][region][country] = {segment: st.number_input(f"Enter the YoY change for {segment} in {country} ({region}) in year {period}:", min_value=-1.000, max_value=2.000, value=0.065, step=0.001, format="%.4f") for segment in segments}

            subsegment_yoy_changes = {}
            for period in previous_periods + future_periods:
                subsegment_yoy_changes[period] = {}
                for region in countries:
                    subsegment_yoy_changes[period][region] = {}
                    for country in countries[region]:
                        subsegment_yoy_changes[period][region][country] = {}
                        for segment in segments:
                            subsegment_yoy_changes[period][region][country][segment] = {subsegment: st.number_input(f"Enter the YoY change for {subsegment} in {segment} ({country}, {region}) in year {period}:", min_value=-1.000, max_value=2.000, value=0.065, step=0.001, format="%.4f") for subsegment in subsegments[segment]}

            # Submit button for form
            submit_button = st.form_submit_button(label="Submit")

            if submit_button:
                st.session_state.data = {
                    'base_year': base_year,
                    'base_value': base_value,
                    'previous_periods': previous_periods,
                    'previous_yoy_changes': previous_yoy_changes,
                    'future_periods': future_periods,
                    'future_yoy_changes': future_yoy_changes,
                    'regions': regions,
                    'region_shares': region_shares,
                    'region_yoy_changes': region_yoy_changes,
                    'countries': countries,
                    'country_shares': country_shares,
                    'country_yoy_changes': country_yoy_changes,
                    'segments': segments,
                    'segment_shares': segment_shares,
                    'segment_yoy_changes': segment_yoy_changes,
                    'subsegments': subsegments,
                    'subsegment_shares': subsegment_shares,
                    'subsegment_yoy_changes': subsegment_yoy_changes
                }
                st.success("Data successfully submitted!")

    elif page == "Review Data":
        st.header("Review Data")

        if 'data' in st.session_state:
            data = st.session_state.data

            base_model_data = process_country_data(data['base_year'], data['base_value'], data['previous_periods'], data['previous_yoy_changes'], is_future=False) + [(data['base_year'], data['base_value'])] + process_country_data(data['base_year'], data['base_value'], data['future_periods'], data['future_yoy_changes'], is_future=True)
            region_data = calculate_region_data(base_model_data, data['region_shares'], data['region_yoy_changes'], data['base_year'], data['base_value'])
            country_data = calculate_country_data(region_data, data['country_shares'], data['country_yoy_changes'], data['base_year'])
            segment_data = calculate_segment_data(country_data, data['segment_shares'], data['segment_yoy_changes'], data['base_year'])
            subsegment_data = calculate_subsegment_data(segment_data, data['subsegments'], data['subsegment_shares'], data['subsegment_yoy_changes'], data['base_year'])

            st.header("Base Model Data")
            st.write(pd.DataFrame(base_model_data, columns=['Year', 'Value']))

            st.header("Region Data")
            st.write(pd.DataFrame(region_data))

            st.header("Country Data")
            st.write(pd.DataFrame(country_data))

            st.header("Segment Data")
            st.write(pd.DataFrame(segment_data))

            st.header("Subsegment Data")
            st.write(pd.DataFrame(subsegment_data))

            output = save_data_to_excel(
                base_model_data, region_data, country_data, segment_data, subsegment_data
            )
            
            if output:
                st.download_button(
                    label="Download Excel File",
                    data=output.getvalue(),
                    file_name="Global Market Data.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )
if __name__ == "__main__":
    password = st.sidebar.text_input("Password", type="password")

    # Check password
    if password == PASSWORD:
        main()  # If password is correct, run the app
    else:
        st.error("Incorrect password. Please try again.")
